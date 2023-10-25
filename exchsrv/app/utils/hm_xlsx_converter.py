#!python
#cython: language_level=3, always_allow_keywords=True
from .num2text import decimal2text, int_units, exp_units

from copy import copy
from calendar import monthrange
from decimal import Decimal
from zipfile import ZipFile
from pathlib import Path
from flask import current_app
from datetime import datetime
from re import escape, split
from os import remove
from os.path import join, basename, exists
import xml.etree.ElementTree as et
from openpyxl import Workbook, load_workbook
from openpyxl.styles.borders import Border, Side

HEADER_ROWS = 8
HEADER_COLUMNS = 20
FOOTER_ROWS = 12
thin_border_all = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))


def v002_title_converter(word):
    if word.endswith('ии'):
        word = word[:-2] + 'ия'
    elif word.endswith('ой'):
        word = word[:-2] + 'ая'
    elif word.endswith('ому'):
        word = word[:-3] + 'ое'
    elif word.endswith('лу'):
        word = word[:-2] + 'ло'
    elif word.endswith('жу'):
        word = word[:-2] + 'ж'
    elif word.endswith('е'):
        word = word[:-1] + 'а'
    elif word.endswith('не'):
        word = word[:-2] + 'на'
    elif word.endswith('ию'):
        word = word[:-2] + 'ие'
    elif word.endswith('щи'):
        word = word[:-2] + 'щь'
    elif word.endswith('ке'):
        word = word[:-2] + 'ка'
    else:
        word = word
    return word


def v002_string_converter(input_string):
    title = input_string
    try:
        delimiters = ' ', '(', ')', '-'
        pattern = '|'.join(map(escape, delimiters))
        temp_lst = split(pattern, input_string)
        for word in temp_lst:
            title = title.replace(word, v002_title_converter(word), 1)
    except Exception as err:
        current_app.logger.critical(f'IN FUNCTION v002_string_converter() ERROR: {str(err)}')
    return title.capitalize()


def med_direction(code):
    v002_title = 'Неизвестное направление мед.помощи'
    try:
        nsi_v002_file = join(join(current_app.root_path, 'opt', 'nsi'), 'V002.xml')
        tree_v002 = et.parse(nsi_v002_file)
        tree_v002_root = tree_v002.getroot()
        for v002_item in tree_v002_root.findall('zap'):
            if v002_item.find('IDPR') is not None:
                if str(code) == str(v002_item.find('IDPR').text):
                    v002_title = v002_string_converter(str(v002_item.find('PRNAME').text))
            else:
                v002_title = 'Ошибка справочника V002'
    except Exception as err:
        current_app.logger.critical(f'IN FUNCTION med_direction() ERROR: {str(err)}')
    return v002_title


def xml_hm_parser(file_name_bill, file_name_pdn, template_dir, work_dir):
    try:
        attach_list = []
        tree_bill = et.parse(file_name_bill)
        tree_bill_root = tree_bill.getroot()
        tree_pdn = et.parse(file_name_pdn)
        tree_pdn_root = tree_pdn.getroot()
        bill_filename = f'РЕЕСТР_{tree_bill_root.find("SCHET/NSCHET").text}.xlsx' if tree_bill_root.find(
            'SCHET/NSCHET') is not None else 'РЕЕСТР.xlsx'
        bill_sd_z = tree_bill_root.find('ZGLV/SD_Z').text if tree_bill_root.find('ZGLV/SD_Z') is not None else ''
        bill_nschet = tree_bill_root.find('SCHET/NSCHET').text if tree_bill_root.find(
            'SCHET/NSCHET') is not None else ''
        bill_dschet = datetime.fromisoformat(tree_bill_root.find('SCHET/DSCHET').text).strftime(
            '%d.%m.%Y') if tree_bill_root.find('SCHET/DSCHET') is not None else ''
        bill_summav = tree_bill_root.find('SCHET/SUMMAV').text if tree_bill_root.find(
            'SCHET/SUMMAV') is not None else ''
        bill_year = tree_bill_root.find('SCHET/YEAR').text if tree_bill_root.find('SCHET/YEAR') is not None else ''
        bill_month = tree_bill_root.find('SCHET/MONTH').text if tree_bill_root.find('SCHET/MONTH') is not None else ''
        bill_month_str = str(bill_month) if int(bill_month) >= 10 else '0' + str(bill_month)
        bill_daterange = f"""01.{bill_month_str}.{bill_year}-{monthrange(int(bill_year), int(bill_month))[1]}.{bill_month_str}.{bill_year}"""
        bill_plat = tree_bill_root.find('SCHET/PLAT').text if tree_bill_root.find('SCHET/PLAT') is not None else ''
        if bill_plat == '47':
            bill_plat_short = r"""За пределами субъекта Российской Федерации, 
                                  на территории которого выдан полис обязательного медицинского страхования"""
            bill_plat_long = 'ТФОМС ЛО'
        elif bill_plat == '47042':
            bill_plat_short = 'АО "СТРАХОВАЯ КОМПАНИЯ "СОГАЗ-МЕД"'
            bill_plat_long = r"""АО "СТРАХОВАЯ КОМПАНИЯ "СОГАЗ-МЕД" 
                                 ИНН 7728170427 КПП 770801001 
                                 197136, Санкт-Петербург г, Внутригородская территория муниципальный округ округ Петровский, 
                                 Лахтинская ул, дом № 16, литер А"""
        elif bill_plat == '47043':
            bill_plat_short = 'ООО "КАПИТАЛ МС"'
            bill_plat_long = r"""ООО "КАПИТАЛ МС" 
                                 ИНН 7813171100 КПП 770501001 
                                 194044, Санкт-Петербург г, Внутригородская территория муниципальный округ округ Петровский, 
                                 Малый П.С. пр-кт, дом № 7, литер Б"""
        elif bill_plat == '47045':
            bill_plat_short = 'ООО "СМК РЕСО-МЕД"'
            bill_plat_long = r"""ООО "СМК РЕСО-МЕД" 
                                ИНН 5035000265 КПП 503501001 
                                197227, Санкт-Петербург г, Внутригородская территория муниципальный округ Озеро Долгое, 
                                Гаккелевская ул, дом № 21, литер А"""
        else:
            bill_plat_short = 'Не определена'
            bill_plat_long = 'Не определен'
        for bill_item in tree_bill_root.findall('ZAP/PACIENT'):
            bill_item_pacient_fio = et.SubElement(bill_item, 'PACIENT_FIO')
            bill_item_pacient_dr = et.SubElement(bill_item, 'PACIENT_DR')
            bill_item_pacient_w = et.SubElement(bill_item, 'PACIENT_W')
            bill_item_pacient_doc = et.SubElement(bill_item, 'PACIENT_DOC')
            bill_item_pacient_snils = et.SubElement(bill_item, 'PACIENT_SNILS')
            bill_id_pac = bill_item.find('ID_PAC')
            for pdn_item in tree_pdn_root.findall('PERS'):
                pdn_id_pac = pdn_item.find('ID_PAC')
                if bill_id_pac.text == pdn_id_pac.text:
                    pdn_fam = pdn_item.find('FAM').text if pdn_item.find('FAM') is not None else ''
                    pdn_im = pdn_item.find('IM').text if pdn_item.find('IM') is not None else ''
                    pdn_ot = pdn_item.find('OT').text if pdn_item.find('OT') is not None else ''
                    if pdn_fam == '':
                        pdn_fam = pdn_item.find('FAM_P').text if pdn_item.find('FAM_P') is not None else ''
                    if pdn_im == '':
                        pdn_im = pdn_item.find('IM_P').text if pdn_item.find('IM_P') is not None else ''
                    if pdn_ot == '':
                        pdn_ot = pdn_item.find('OT_P').text if pdn_item.find('OT_P') is not None else ''
                    pdn_dr = datetime.fromisoformat(pdn_item.find('DR').text).strftime('%d.%m.%Y') if pdn_item.find(
                        'DR') is not None else ''
                    pdn_w = pdn_item.find('W').text if pdn_item.find('W') is not None else ''
                    pdn_docser = pdn_item.find('DOCSER').text if pdn_item.find('DOCSER') is not None else ''
                    pdn_docnum = pdn_item.find('DOCNUM').text if pdn_item.find('DOCNUM') is not None else ''
                    pdn_snils = pdn_item.find('SNILS').text if pdn_item.find('SNILS') is not None else ''
                    bill_item_pacient_fio.text = pdn_fam + ' ' + pdn_im + ' ' + pdn_ot
                    bill_item_pacient_dr.text = pdn_dr
                    bill_item_pacient_w.text = pdn_w
                    bill_item_pacient_doc.text = pdn_docser + ' ' + pdn_docnum
                    bill_item_pacient_snils.text = pdn_snils
        book = Workbook()
        sheet = book.active
        filename_header = join(template_dir, 'proto_header.xlsx')
        wb_header = load_workbook(filename_header)
        ws_header = wb_header.worksheets[0]
        for rw in range(1, HEADER_ROWS):
            for cl in range(1, HEADER_COLUMNS):
                current_cell = ws_header.cell(row=rw, column=cl)
                sheet.cell(row=rw, column=cl).value = current_cell.value
                if current_cell.has_style:
                    sheet.cell(row=rw, column=cl).font = copy(current_cell.font)
                    sheet.cell(row=rw, column=cl).border = copy(current_cell.border)
                    sheet.cell(row=rw, column=cl).fill = copy(current_cell.fill)
                    sheet.cell(row=rw, column=cl).number_format = copy(current_cell.number_format)
                    sheet.cell(row=rw, column=cl).protection = copy(current_cell.protection)
                    sheet.cell(row=rw, column=cl).alignment = copy(current_cell.alignment)
        for rw in range(1, HEADER_ROWS):
            for cl in range(1, HEADER_COLUMNS):
                if sheet.cell(row=rw, column=cl).value == '{###BILL_REESTR###}':
                    sheet.cell(row=rw, column=cl).value = f'РЕЕСТР СЧЕТА №{bill_nschet} от {bill_dschet}'
                if sheet.cell(row=rw, column=cl).value == '{###INS_ORGANISATION###}':
                    sheet.cell(row=rw, column=cl).value = f'{bill_plat_short}'
                if sheet.cell(row=rw, column=cl).value == '{###PERIOD###}':
                    sheet.cell(row=rw, column=cl).value = f'период формирования: {bill_daterange}'
        row_counter = 0
        for zap_item in tree_bill_root.findall('ZAP'):
            if zap_item:
                zap_id = zap_item.find('N_ZAP').text if zap_item.find('N_ZAP') is not None else ''
                zap_patient_fio = zap_item.find('PACIENT/PACIENT_FIO').text if zap_item.find(
                                                                            'PACIENT/PACIENT_FIO') is not None else ''
                zap_patient_w = zap_item.find('PACIENT/PACIENT_W').text if zap_item.find(
                                                                            'PACIENT/PACIENT_W') is not None else ''
                zap_patient_dr = zap_item.find('PACIENT/PACIENT_DR').text if zap_item.find(
                                                                            'PACIENT/PACIENT_DR') is not None else ''
                zap_patient_doc = zap_item.find('PACIENT/PACIENT_DOC').text if zap_item.find(
                                                                            'PACIENT/PACIENT_DOC') is not None else ''
                zap_patient_snils = zap_item.find('PACIENT/PACIENT_SNILS').text if zap_item.find(
                                                                            'PACIENT/PACIENT_SNILS') is not None else ''
                zap_patient_npolis = zap_item.find('PACIENT/NPOLIS').text if zap_item.find(
                                                                            'PACIENT/NPOLIS') is not None else ''
                zap_patient_vidpom = zap_item.find('Z_SL/VIDPOM').text if zap_item.find(
                                                                            'Z_SL/VIDPOM') is not None else ''
                zap_patient_ds1 = zap_item.find('Z_SL/SL/DS1').text if zap_item.find('Z_SL/SL/DS1') is not None else ''
                zap_patient_date1 = datetime.fromisoformat(zap_item.find('Z_SL/SL/DATE_1').text).strftime(
                                                    '%d.%m.%Y') if zap_item.find('Z_SL/SL/DATE_1') is not None else ''
                zap_patient_date2 = datetime.fromisoformat(zap_item.find('Z_SL/SL/DATE_2').text).strftime(
                                                    '%d.%m.%Y') if zap_item.find('Z_SL/SL/DATE_2') is not None else ''
                zap_patient_date12 = zap_patient_date1 + '-' + zap_patient_date2
                zap_patient_edcol = zap_item.find('Z_SL/SL/ED_COL').text if zap_item.find(
                                                                                'Z_SL/SL/ED_COL') is not None else ''
                zap_patient_profil = zap_item.find('Z_SL/SL/PROFIL').text if zap_item.find(
                                                                                'Z_SL/SL/PROFIL') is not None else ''
                zap_patient_prvs = zap_item.find('Z_SL/SL/PRVS').text if zap_item.find(
                                                                                 'Z_SL/SL/PRVS') is not None else ''
                zap_patient_summ = zap_item.find('Z_SL/SL/SUM_M').text if zap_item.find(
                                                                                 'Z_SL/SL/SUM_M') is not None else ''
                zap_patient_rslt = zap_item.find('Z_SL/RSLT').text if zap_item.find('Z_SL/RSLT') is not None else ''
                zap_row = [zap_id, zap_patient_fio, zap_patient_w, zap_patient_dr, '', zap_patient_doc, '', '',
                           zap_patient_snils, zap_patient_npolis, zap_patient_vidpom, zap_patient_ds1,
                           zap_patient_date12, zap_patient_edcol, zap_patient_profil, zap_patient_prvs,
                           zap_patient_summ, zap_patient_summ, zap_patient_rslt]
                for col, value in enumerate(zap_row):
                    sheet.cell(row=row_counter + HEADER_ROWS, column=col + 1).value = value
                    sheet.cell(row=row_counter + HEADER_ROWS, column=col + 1).border = thin_border_all
                row_counter += 1
        sheet.cell(row=row_counter + HEADER_ROWS + 1, column=2).value = f'Итого: {decimal2text(Decimal(bill_summav), int_units=int_units, exp_units=exp_units)}'
        sheet.cell(row=row_counter + HEADER_ROWS + 1, column=17).value = f'Всего услуг: {bill_sd_z}'
        filename_footer = join(template_dir, 'proto_footer.xlsx')
        wb_footer = load_workbook(filename_footer)
        ws_footer = wb_footer.worksheets[0]
        for rw in range(1, FOOTER_ROWS):
            for cl in range(1, HEADER_COLUMNS):
                current_cell = ws_footer.cell(row=rw, column=cl)
                sheet.cell(row=row_counter + HEADER_ROWS + 3 + rw, column=cl).value = current_cell.value
                if sheet.cell(row=row_counter + HEADER_ROWS + 3 + rw, column=cl).value == '{###BILL_DATE###}':
                    sheet.cell(row=row_counter + HEADER_ROWS + 3 + rw, column=cl).value = f'Дата {bill_dschet}'
        reestr_file_name = join(work_dir, bill_filename)
        attach_list.append(reestr_file_name)
        book.save(reestr_file_name)
        book = Workbook()
        sheet = book.active
        bill_filename = f'СЧЕТ_{tree_bill_root.find("SCHET/NSCHET").text}.xlsx' if tree_bill_root.find(
                                                                        'SCHET/NSCHET') is not None else 'СЧЕТ.xlsx'
        header_second_start = 1
        filename_header_second = join(template_dir, 'proto_header_second.xlsx')
        wb_header_second = load_workbook(filename_header_second)
        ws_header_second = wb_header_second.worksheets[0]
        for rw in range(1, FOOTER_ROWS + 2):
            for cl in range(1, HEADER_COLUMNS):
                current_cell = ws_header_second.cell(row=rw, column=cl)
                sheet.cell(row=header_second_start + rw, column=cl).value = current_cell.value
                if sheet.cell(row=header_second_start + rw, column=cl).value == '{##PLAT###}':
                    sheet.cell(row=header_second_start + rw, column=cl).value = f'Плательщик: {bill_plat_long}'
                if sheet.cell(row=header_second_start + rw, column=cl).value == '{###BILL###}':
                    sheet.cell(row=header_second_start + rw, column=cl).value = f'СЧЕТ {bill_nschet} от {bill_dschet}'
                if sheet.cell(row=header_second_start + rw, column=cl).value == '{###PERIOD###}':
                    sheet.cell(row=header_second_start + rw, column=cl).value = f'за период {bill_daterange}'
                if current_cell.has_style:
                    sheet.cell(row=header_second_start + rw, column=cl).font = copy(current_cell.font)
                    sheet.cell(row=header_second_start + rw, column=cl).border = copy(current_cell.border)
                    sheet.cell(row=header_second_start + rw, column=cl).fill = copy(current_cell.fill)
                    sheet.cell(row=header_second_start + rw, column=cl).number_format = copy(current_cell.number_format)
                    sheet.cell(row=header_second_start + rw, column=cl).protection = copy(current_cell.protection)
                    sheet.cell(row=header_second_start + rw, column=cl).alignment = copy(current_cell.alignment)
        footer_second_start = header_second_start + FOOTER_ROWS
        med_profile_dict = {}
        med_profile_sum = {}
        for zap_item in tree_bill_root.findall('ZAP'):
            if zap_item:
                zap_patient_profil = zap_item.find('Z_SL/SL/PROFIL').text if zap_item.find(
                                                                                'Z_SL/SL/PROFIL') is not None else ''
                zap_patient_summ = float(zap_item.find('Z_SL/SL/SUM_M').text) if zap_item.find(
                                                                                'Z_SL/SL/SUM_M') is not None else 0.0
                med_profile_dict[zap_patient_profil] = int(med_profile_dict.get(zap_patient_profil, 0)) + 1
                med_profile_sum[zap_patient_profil] = float(
                    med_profile_sum.get(zap_patient_profil, 0.0)) + zap_patient_summ
        for profile, summ in med_profile_sum.items():
            med_profile_sum[profile] = [med_profile_dict.get(profile, 0), med_profile_sum[profile]]
        rows_len = len(med_profile_sum.items())
        bill_sum = 0
        bill_count = 0
        for rw in range(1, rows_len):
            sheet.cell(row=footer_second_start + rw, column=2).value = list(med_profile_sum.items())[rw - 1][0]
            sheet.cell(row=footer_second_start + rw, column=2).border = thin_border_all
            sheet.cell(row=footer_second_start + rw, column=3).value = med_direction(
                list(med_profile_sum.items())[rw - 1][0])
            sheet.cell(row=footer_second_start + rw, column=3).border = thin_border_all
            sheet.cell(row=footer_second_start + rw, column=4).value = list(med_profile_sum.items())[rw - 1][1][0]
            bill_count = bill_count + int(list(med_profile_sum.items())[rw - 1][1][0])
            sheet.cell(row=footer_second_start + rw, column=4).border = thin_border_all
            sheet.cell(row=footer_second_start + rw, column=5).value = str(
                list(med_profile_sum.items())[rw - 1][1][0]) + '.00'
            sheet.cell(row=footer_second_start + rw, column=5).border = thin_border_all
            sheet.cell(row=footer_second_start + rw, column=6).value = list(med_profile_sum.items())[rw - 1][1][1]
            bill_sum = bill_sum + float(list(med_profile_sum.items())[rw - 1][1][1])
            sheet.cell(row=footer_second_start + rw, column=6).border = thin_border_all
        sheet.cell(row=footer_second_start + rows_len, column=2).border = thin_border_all
        sheet.cell(row=footer_second_start + rows_len, column=3).value = 'ИТОГО по счету'
        sheet.cell(row=footer_second_start + rows_len, column=3).border = thin_border_all
        sheet.cell(row=footer_second_start + rows_len, column=4).value = bill_count
        sheet.cell(row=footer_second_start + rows_len, column=4).border = thin_border_all
        sheet.cell(row=footer_second_start + rows_len, column=5).value = str(bill_count) + '.00'
        sheet.cell(row=footer_second_start + rows_len, column=5).border = thin_border_all
        sheet.cell(row=footer_second_start + rows_len, column=6).value = bill_sum
        sheet.cell(row=footer_second_start + rows_len, column=6).border = thin_border_all
        filename_footer_second = join(template_dir, 'proto_footer_second.xlsx')
        wb_footer_second = load_workbook(filename_footer_second)
        ws_footer_second = wb_footer_second.worksheets[0]
        for rw in range(1, 6):
            for cl in range(1, HEADER_COLUMNS):
                current_cell = ws_footer_second.cell(row=rw, column=cl)
                sheet.cell(row=footer_second_start + rows_len + 1 + rw, column=cl).value = current_cell.value
                if sheet.cell(row=footer_second_start + rows_len + 1 + rw, column=cl).value == '{###TOTAL_SUM_TEXT###}':
                    sheet.cell(row=footer_second_start + rows_len + 1 + rw,
                               column=cl).value = f'Сумма прописью: {decimal2text(Decimal(bill_sum), int_units=int_units, exp_units=exp_units)}'
                if current_cell.has_style:
                    sheet.cell(row=footer_second_start + rows_len + 1 + rw, column=cl).font = copy(current_cell.font)
                    sheet.cell(row=footer_second_start + rows_len + 1 + rw, column=cl).border = copy(
                        current_cell.border)
                    sheet.cell(row=footer_second_start + rows_len + 1 + rw, column=cl).fill = copy(current_cell.fill)
                    sheet.cell(row=footer_second_start + rows_len + 1 + rw, column=cl).number_format = copy(
                        current_cell.number_format)
                    sheet.cell(row=footer_second_start + rows_len + 1 + rw, column=cl).protection = copy(
                        current_cell.protection)
                    sheet.cell(row=footer_second_start + rows_len + 1 + rw, column=cl).alignment = copy(
                        current_cell.alignment)
        schet_file_name = join(work_dir, bill_filename)
        attach_list.append(schet_file_name)
        book.save(schet_file_name)
        zip_file_name = join(work_dir, f'РЕЕСТР_СЧЕТ_{tree_bill_root.find("SCHET/NSCHET").text}.zip')
        with ZipFile(zip_file_name, 'w') as zipObj:
            for file in attach_list:
                zipObj.write(file, arcname=basename(file))
        for file in attach_list:
            if exists(file):
                remove(file)
        return Path(zip_file_name).name
    except Exception as err:
        current_app.logger.critical(f'IN FUNCTION xml_hm_parser() ERROR: {str(err)}')
